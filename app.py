# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, and_, func, TypeDecorator, String, Enum as SQLAlchemyEnum
from sqlalchemy.orm import selectinload
from datetime import date, datetime, timedelta
import os
import openpyxl
import json
import math
import pytz
from io import BytesIO
import uuid
import calendar
import secrets
import logging # Loggingを追加

# --- .envファイルを読み込む ---
from dotenv import load_dotenv
load_dotenv()

# --- ログイン機能のためのインポート ---
from flask_login import LoginManager, current_user, login_user, logout_user, login_required, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash

# --- スプレッドシート連携用 ---
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- 1. アプリの初期化と設定 ---
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "a-very-secret-key-for-local-development")

# --- Logging設定 ---
logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

db_url = os.environ.get('DATABASE_URL', 'sqlite:///instance/tasks.db')
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

if 'sqlite' in db_url:
    instance_path = os.path.join(app.root_path, 'instance')
    os.makedirs(instance_path, exist_ok=True)
    db_url = f'sqlite:///{os.path.join(instance_path, "tasks.db")}'

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

engine_options = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}

db = SQLAlchemy(app, engine_options=engine_options)

class DateAsString(TypeDecorator):
    impl = String
    cache_ok = True
    def process_bind_param(self, value, dialect): return value.isoformat() if value is not None else None
    def process_result_value(self, value, dialect): return date.fromisoformat(value) if value is not None else None

# --- 2. データベースモデルの定義 ---
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)
    password_reset_required = db.Column(db.Boolean, default=False, nullable=False)
    spreadsheet_url = db.Column(db.String(255), nullable=True)
    master_tasks = db.relationship('MasterTask', backref='user', lazy=True, cascade="all, delete-orphan")
    summaries = db.relationship('DailySummary', backref='user', lazy=True, cascade="all, delete-orphan")
    task_templates = db.relationship('TaskTemplate', backref='user', lazy=True, cascade="all, delete-orphan")

# --- 繰り返しタイプ用のEnum ---
class RecurrenceType(SQLAlchemyEnum):
    NONE = 'none'
    DAILY = 'daily'
    WEEKLY = 'weekly'

class MasterTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(100), nullable=False)
    due_date = db.Column(DateAsString, default=lambda: get_jst_today(), nullable=False) # 通常タスクの期限日 or 繰り返しタスクの開始日
    is_urgent = db.Column(db.Boolean, default=False, nullable=False)
    is_habit = db.Column(db.Boolean, default=False, nullable=False) # 習慣フラグ
    recurrence_type = db.Column(db.Enum('none', 'daily', 'weekly', name='recurrence_type_enum'), default='none', nullable=False) # 繰り返し種別
    recurrence_days = db.Column(db.String(7), nullable=True) # 繰り返し曜日 (例: '01234') 月曜=0
    last_reset_date = db.Column(DateAsString, nullable=True) # 最後に完了状態がリセットされた日
    subtasks = db.relationship('SubTask', backref='master_task', lazy=True, cascade="all, delete-orphan")

class SubTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    master_id = db.Column(db.Integer, db.ForeignKey('master_task.id'), nullable=False)
    content = db.Column(db.String(100), nullable=False)
    grid_count = db.Column(db.Integer, default=1, nullable=False)
    is_completed = db.Column(db.Boolean, default=False)
    completion_date = db.Column(DateAsString, nullable=True)

class DailySummary(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    summary_date = db.Column(DateAsString, nullable=False)
    streak = db.Column(db.Integer, default=0)
    average_grids = db.Column(db.Float, default=0.0)

class TaskTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(100), nullable=False)
    subtask_templates = db.relationship('SubtaskTemplate', backref='task_template', lazy=True, cascade="all, delete-orphan")

class SubtaskTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey('task_template.id'), nullable=False)
    content = db.Column(db.String(100), nullable=False)
    grid_count = db.Column(db.Integer, default=1, nullable=False)

# --- 3. ログイン管理 ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.before_request
def require_password_change():
    if current_user.is_authenticated and current_user.password_reset_required:
        # --- PWA/オフライン同期用のエンドポイントを許可リストに追加 ---
        # ★習慣カレンダー関連のルートを追加
        allowed_endpoints = ['settings', 'logout', 'static', 'health_check', 'serve_sw', 'serve_manifest', 'sync_api', 'habit_calendar', 'habit_calendar_data']
        if request.endpoint not in allowed_endpoints:
            flash('セキュリティのため、新しいパスワードを設定してください。', 'warning')
            return redirect(url_for('settings'))

# --- Helper Functions ---
def get_jst_today():
    """Get today's date in JST timezone."""
    return datetime.now(pytz.timezone('Asia/Tokyo')).date()

# --- 繰り返しタスクのリセット処理 ---
def reset_recurring_tasks_if_needed(user_id):
    today = get_jst_today()
    tasks_to_reset = MasterTask.query.filter(
        MasterTask.user_id == user_id,
        MasterTask.recurrence_type != 'none',
        # 今日より前にリセットされているか、まだ一度もリセットされていないタスク
        or_(MasterTask.last_reset_date == None, MasterTask.last_reset_date < today)
    ).all()

    reset_count = 0
    for task in tasks_to_reset:
        should_reset_today = False
        # 開始日 (due_date) 以降でなければリセットしない
        if task.due_date > today:
            continue

        if task.recurrence_type == 'daily':
            should_reset_today = True
        elif task.recurrence_type == 'weekly' and task.recurrence_days:
            today_weekday = str(today.weekday()) # Monday is 0 and Sunday is 6
            if today_weekday in task.recurrence_days:
                should_reset_today = True

        if should_reset_today:
            # サブタスクの完了状態と完了日をリセット
            subtasks_updated = SubTask.query.filter(
                SubTask.master_id == task.id,
                SubTask.is_completed == True
            ).update({
                SubTask.is_completed: False,
                SubTask.completion_date: None
            }, synchronize_session=False)

            if subtasks_updated > 0:
                reset_count += subtasks_updated # リセットされたサブタスク数をカウント

            task.last_reset_date = today # 最終リセット日を今日に更新

    if reset_count > 0:
        db.session.commit()
        app.logger.info(f"User {user_id}: Reset {reset_count} subtasks for recurring master tasks for {today}.")
    elif tasks_to_reset: # リセット対象があったが曜日の関係などでリセットされなかった場合も last_reset_date の更新をコミット
        db.session.commit()


def update_summary(user_id):
    """Update daily summary (streak, average grids) for the user."""
    today = get_jst_today()
    # 習慣タスクのみ or 全タスクで計算するか要検討 (今は全タスク)
    grids_by_date = db.session.query(
        SubTask.completion_date, func.sum(SubTask.grid_count)
    ).join(MasterTask).filter(
        MasterTask.user_id == user_id,
        SubTask.is_completed == True,
        SubTask.completion_date != None # 完了日がないものは除外
    ).group_by(SubTask.completion_date).order_by(SubTask.completion_date.desc()).limit(30).all() # 直近30日分

    average_grids = sum(g[1] for g in grids_by_date) / len(grids_by_date) if grids_by_date else 0.0

    # ストリーク計算 (完了日が存在するSubTaskを基に)
    completed_dates = db.session.query(
        SubTask.completion_date
    ).join(MasterTask).filter(
        MasterTask.user_id == user_id,
        SubTask.is_completed == True,
        SubTask.completion_date != None
    ).distinct().all()

    streak = 0
    if completed_dates:
        unique_dates_set = {d[0] for d in completed_dates}
        check_date = today
        # 今日または昨日完了していればストリーク開始
        if today in unique_dates_set or (today - timedelta(days=1)) in unique_dates_set:
             # もし今日完了していなければ、昨日から遡る
            if today not in unique_dates_set:
                check_date = today - timedelta(days=1)
            # 完了している限り遡る
            while check_date in unique_dates_set:
                streak += 1
                check_date -= timedelta(days=1)

    summary = DailySummary.query.filter_by(user_id=user_id, summary_date=today).first()
    if not summary:
        summary = DailySummary(user_id=user_id, summary_date=today)
        db.session.add(summary)

    summary.streak = streak
    summary.average_grids = round(average_grids, 2)
    db.session.commit()


def cleanup_old_tasks(user_id):
    # This function is not called automatically anymore, consider calling it periodically or manually if needed.
    cleanup_threshold = get_jst_today() - timedelta(days=32)
    # Delete non-recurring completed tasks older than threshold
    old_subtasks_query = SubTask.query.join(MasterTask).filter(
        MasterTask.user_id == user_id,
        MasterTask.recurrence_type == 'none', # Only non-recurring
        SubTask.is_completed == True,
        SubTask.completion_date < cleanup_threshold
    )
    # Also delete master tasks if all their subtasks are old and completed
    master_ids_to_check = [st.master_id for st in old_subtasks_query.all()]

    deleted_subtask_count = old_subtasks_query.delete(synchronize_session=False)

    deleted_master_count = 0
    if master_ids_to_check:
        masters_to_delete = MasterTask.query.filter(
            MasterTask.id.in_(master_ids_to_check),
            ~MasterTask.subtasks.any(or_(SubTask.is_completed == False, SubTask.completion_date >= cleanup_threshold))
        )
        deleted_master_count = masters_to_delete.delete(synchronize_session=False)

    if deleted_subtask_count > 0 or deleted_master_count > 0:
        db.session.commit()
        app.logger.info(f"User {user_id}: Cleaned up {deleted_master_count} master tasks and {deleted_subtask_count} old non-recurring subtasks.")


# --- 4. 【重要】手動データベース初期化ルート ---
# (安全のため、環境変数等で有効/無効を切り替えられるようにすることを推奨)
@app.route('/init-db/<secret_key>')
def init_db(secret_key):
    # シークレットキーの比較 (環境変数から取得)
    expected_key = os.environ.get("FLASK_SECRET_KEY")
    if not expected_key or secret_key != expected_key:
        app.logger.warning("Invalid secret key used for DB initialization attempt.")
        return "認証キーが正しくありません。", 403

    try:
        with app.app_context():
            app.logger.info("Initializing database...")
            db.create_all() # This will add new columns if they don't exist
            app.logger.info("Database tables checked/created/updated.")

            admin_username = os.environ.get('ADMIN_USERNAME')
            if admin_username:
                admin_user = User.query.filter_by(username=admin_username).first()
                if admin_user:
                    if not admin_user.is_admin:
                        admin_user.is_admin = True
                        db.session.commit()
                        app.logger.info(f"User '{admin_username}' set as admin.")
                        return f"データベースが初期化/更新され、ユーザー '{admin_username}' が管理者に設定されました。"
                    else:
                        app.logger.info(f"User '{admin_username}' is already an admin.")
                        return f"データベースは初期化/更新済みで、ユーザー '{admin_username}' は既に管理者です。"
                else:
                    app.logger.info(f"Admin user '{admin_username}' not found. Please register first.")
                    return f"データベースは初期化/更新されましたが、管理者ユーザー '{admin_username}' はまだ登録されていません。先にその名前でユーザー登録してから、再度このURLにアクセスしてください。"
            else:
                app.logger.info("ADMIN_USERNAME not set in environment.")
                return "データベースが初期化/更新されました (管理者設定なし)。"
    except Exception as e:
        app.logger.error(f"Error during database initialization/update: {e}", exc_info=True)
        return f"データベース初期化/更新中にエラーが発生しました: {e}", 500


# --- 6. 認証・ログイン関連のルート ---
@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("todo_list"))
    else:
        return redirect(url_for("login"))

@app.route("/healthz")
def health_check():
    # Basic check: just return OK
    # More advanced checks could involve trying a DB connection
    return "OK", 200

# --- PWA対応: sw.js と manifest.json を配信するルート ---
@app.route('/sw.js')
def serve_sw():
    # ルートディレクトリの sw.js を配信
    return send_file('sw.js', mimetype='application/javascript')

@app.route('/manifest.json')
def serve_manifest():
    # static ディレクトリの manifest.json を配信
    return send_file('static/manifest.json', mimetype='application/manifest+json')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('todo_list'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if not username or not password:
            flash('ユーザー名とパスワードを入力してください。')
            return redirect(url_for('register'))
        user = User.query.filter_by(username=username).first()
        if user:
            flash('このユーザー名は既に使用されています。')
            return redirect(url_for('register'))
        try:
            new_user = User(username=username, password_hash=generate_password_hash(password, method='pbkdf2:sha256'))
            db.session.add(new_user)
            db.session.commit()
            login_user(new_user)
            app.logger.info(f"New user registered: {username}")
            return redirect(url_for('todo_list'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error during registration for {username}: {e}", exc_info=True)
            flash('登録中にエラーが発生しました。')
            return redirect(url_for('register'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('todo_list'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = bool(request.form.get('remember'))
        if not username or not password:
            flash('ユーザー名とパスワードを入力してください。')
            return redirect(url_for('login'))

        user = User.query.filter_by(username=username).first()

        # Admin master password check
        admin_username = os.environ.get('ADMIN_USERNAME')
        admin_password = os.environ.get('ADMIN_PASSWORD')

        # Check for admin master password login first
        if user and user.username == admin_username and admin_password and password == admin_password:
            login_user(user, remember=remember)
            flash('管理者としてマスターパスワードでログインしました。')
            app.logger.info(f"Admin user {username} logged in with master password.")
            return redirect(url_for('todo_list'))

        # Standard login check
        if not user or not check_password_hash(user.password_hash, password):
            flash('ユーザー名またはパスワードが正しくありません。')
            app.logger.warning(f"Failed login attempt for username: {username}")
            return redirect(url_for('login'))

        login_user(user, remember=remember)
        app.logger.info(f"User {username} logged in successfully.")
        return redirect(url_for('todo_list'))

    return render_template('login.html')


@app.route("/logout")
@login_required
def logout():
    app.logger.info(f"User {current_user.username} logging out.")
    logout_user()
    flash('ログアウトしました。')
    return redirect(url_for("login"))

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        try:
            if 'update_url' in request.form:
                url = request.form.get('spreadsheet_url', '').strip()
                # Simple validation (more robust validation might be needed)
                if url and url.startswith('https://docs.google.com/spreadsheets/'):
                    current_user.spreadsheet_url = url
                    db.session.commit()
                    flash('スプレッドシートURLを保存しました。')
                    app.logger.info(f"User {current_user.username} updated spreadsheet URL.")
                else:
                    flash('有効なGoogleスプレッドシートURLを入力してください。')

            elif 'change_password' in request.form:
                current_password = request.form.get('current_password')
                new_password = request.form.get('new_password')
                confirm_password = request.form.get('confirm_password')

                admin_username = os.environ.get('ADMIN_USERNAME')
                admin_password = os.environ.get('ADMIN_PASSWORD')
                is_admin_master_password = (current_user.username == admin_username and admin_password and current_password == admin_password)

                if not current_password or not new_password or not confirm_password:
                    flash('すべてのパスワード欄を入力してください。')
                elif not check_password_hash(current_user.password_hash, current_password) and not is_admin_master_password:
                    flash('現在のパスワードが正しくありません。')
                    app.logger.warning(f"User {current_user.username} failed password change attempt (incorrect current password).")
                elif new_password != confirm_password:
                    flash('新しいパスワードが一致しません。')
                elif len(new_password) < 4: # Add minimum length check?
                    flash('パスワードは4文字以上で設定してください。')
                else:
                    current_user.password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
                    current_user.password_reset_required = False # パスワード変更したのでリセット要求を解除
                    db.session.commit()
                    flash('パスワードが正常に変更されました。')
                    app.logger.info(f"User {current_user.username} successfully changed password.")
                    # If password change was forced, redirect to main page
                    if request.args.get('force_change'):
                        return redirect(url_for('todo_list'))
                    return redirect(url_for('settings')) # 通常は設定ページにリダイレクト

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error in settings POST for user {current_user.username}: {e}", exc_info=True)
            flash('設定の保存中にエラーが発生しました。')

        return redirect(url_for('settings')) # エラーがあっても設定ページにリダイレクト

    # --- ★ GET Request Logic (削除日計算の復活) ---
    days_until_deletion = None
    try:
        cleanup_threshold_days = 32 # 32日
        
        # 最も古い完了済み(かつ非繰り返し)タスクの完了日を取得
        oldest_completed_subtask = SubTask.query.join(MasterTask).filter(
            MasterTask.user_id == current_user.id,
            MasterTask.recurrence_type == 'none', # Only non-recurring
            SubTask.is_completed == True,
            SubTask.completion_date != None
        ).order_by(SubTask.completion_date.asc()).first()

        if oldest_completed_subtask:
            today = get_jst_today()
            oldest_completion_date = oldest_completed_subtask.completion_date
            # 削除予定日を計算
            deletion_date = oldest_completion_date + timedelta(days=cleanup_threshold_days)
            # 今日から削除予定日までの日数を計算
            days_until_deletion = (deletion_date - today).days
        
    except Exception as e:
        app.logger.error(f"Error calculating days_until_deletion for user {current_user.id}: {e}", exc_info=True)
        days_until_deletion = None # エラー時は None のまま
    # --- ★ ロジックここまで ---

    sa_email = os.environ.get('SERVICE_ACCOUNT_EMAIL', '（管理者が設定してください）')
    force_password_change = current_user.password_reset_required

    return render_template('settings.html',
                           sa_email=sa_email,
                           days_until_deletion=days_until_deletion, # ★ 修正
                           force_password_change=force_password_change)


# Serve static files required by templates
@app.route('/static/calendar.js')
def calendar_js():
    return send_file('static/calendar.js', mimetype='application/javascript')

# --- PWA対応: static/offline.js を配信するルート ---
@app.route('/static/offline.js')
def serve_offline_js():
    return send_file('static/offline.js', mimetype='application/javascript')

# --- 6. Todoアプリ本体のルート ---
@app.route('/todo')
@app.route('/todo/<date_str>')
@login_required
def todo_list(date_str=None):
    # cleanup_old_tasks(current_user.id) # 自動クリーンアップは一旦停止
    reset_recurring_tasks_if_needed(current_user.id) # ★繰り返しタスクのリセット処理を実行

    if date_str is None:
        target_date = get_jst_today()
    else:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            app.logger.warning(f"Invalid date format received: {date_str}. Redirecting to today.")
            return redirect(url_for('todo_list'))

    # --- カレンダー用のタスク数取得 ---
    first_day_of_month = target_date.replace(day=1)
    next_month_first_day = (first_day_of_month + timedelta(days=32)).replace(day=1)

    # 未完了の「通常」タスク数を日付ごとにカウント
    uncompleted_tasks_count = db.session.query(
        MasterTask.due_date, func.count(MasterTask.id)
    ).outerjoin(SubTask).filter(
        MasterTask.user_id == current_user.id,
        MasterTask.due_date >= first_day_of_month,
        MasterTask.due_date < next_month_first_day,
        MasterTask.recurrence_type == 'none', # 通常タスクのみ
        MasterTask.subtasks.any(SubTask.is_completed == False) # 未完了のサブタスクを持つもの
    ).group_by(MasterTask.due_date).all()

    task_counts_for_js = {d.isoformat(): c for d, c in uncompleted_tasks_count}

    # --- ★表示するタスクの取得 (通常タスクと繰り返しタスクを分ける) ---
    today_weekday = str(target_date.weekday())

    # 一つのクエリで関連する可能性のあるタスクを全て取得 (効率化)
    all_master_tasks = MasterTask.query.options(
        selectinload(MasterTask.subtasks) # サブタスクも一緒に読み込む
    ).filter(
        MasterTask.user_id == current_user.id,
        MasterTask.subtasks.any() # サブタスクが1つ以上存在する
    ).order_by(MasterTask.is_urgent.desc(), MasterTask.due_date.asc(), MasterTask.id.asc()).all()

    daily_tasks_for_template = []       # 今日の通常タスク
    recurring_tasks_for_template = []   # 今日の繰り返しタスク

    for mt in all_master_tasks:
        is_visible_today = False
        # 繰り返しタスクの判定
        if mt.recurrence_type != 'none' and mt.due_date <= target_date: # 開始日以降
            if mt.recurrence_type == 'daily':
                is_visible_today = True
            elif mt.recurrence_type == 'weekly' and mt.recurrence_days and today_weekday in mt.recurrence_days:
                is_visible_today = True
        # 通常タスクの判定 (期限日が今日)
        elif mt.recurrence_type == 'none' and mt.due_date == target_date:
            is_visible_today = True

        if not is_visible_today:
            continue # 今日表示するタスクでなければスキップ

        # 表示するサブタスクを決定
        if mt.recurrence_type != 'none':
            # 繰り返しタスクは常にすべてのサブタスクを表示対象とする
            visible_subtasks = mt.subtasks
        else:
            # 通常タスクは「未完了」または「今日完了した」サブタスクのみ表示
            visible_subtasks = [st for st in mt.subtasks if not st.is_completed or st.completion_date == target_date]

        # 表示するサブタスクがなければ、このマスタタスクは表示しない
        if not visible_subtasks:
            continue

        mt.visible_subtasks = sorted(visible_subtasks, key=lambda x: x.id) # 表示順をIDでソート
        subtasks_as_dicts = [
            {"id": st.id, "content": st.content, "is_completed": st.is_completed, "grid_count": st.grid_count}
            for st in mt.visible_subtasks
        ]
        mt.visible_subtasks_json = json.dumps(subtasks_as_dicts)

        # 今日の表示対象サブタスクが全て完了しているか
        mt.all_completed_today = all(st.is_completed for st in mt.visible_subtasks) if mt.visible_subtasks else False

        # ヘッダー表示用の最終完了日を計算 (全サブタスクベース)
        all_completed_ever = all(st.is_completed for st in mt.subtasks)
        if all_completed_ever:
            completion_dates = [st.completion_date for st in mt.subtasks if st.completion_date]
            mt.last_completion_date = max(completion_dates) if completion_dates else None
        else:
            mt.last_completion_date = None

        # リストに振り分ける
        if mt.recurrence_type != 'none':
            recurring_tasks_for_template.append(mt)
        else:
            daily_tasks_for_template.append(mt)


    # --- グリッド計算 (今日の表示タスクすべて) ---
    all_subtasks_for_day_grid = []
    for mt in daily_tasks_for_template:
        all_subtasks_for_day_grid.extend(mt.visible_subtasks)
    for mt in recurring_tasks_for_template:
        all_subtasks_for_day_grid.extend(mt.visible_subtasks) # 繰り返しもグリッドに含める

    total_grid_count = sum(sub.grid_count for sub in all_subtasks_for_day_grid)
    completed_grid_count = sum(sub.grid_count for sub in all_subtasks_for_day_grid if sub.is_completed)
    GRID_COLS, base_rows = 10, 2
    required_rows = math.ceil(total_grid_count / GRID_COLS) if total_grid_count > 0 else 1
    grid_rows = max(base_rows, required_rows)

    update_summary(current_user.id) # サマリーを更新
    latest_summary = DailySummary.query.filter(DailySummary.user_id == current_user.id).order_by(DailySummary.summary_date.desc()).first()

    return render_template(
        'index.html',
        daily_tasks=daily_tasks_for_template, # ★通常タスクを渡す
        recurring_tasks=recurring_tasks_for_template, # ★繰り返しタスクを渡す
        current_date=target_date,
        today=get_jst_today(),
        total_grid_count=total_grid_count,
        completed_grid_count=completed_grid_count,
        GRID_COLS=GRID_COLS,
        grid_rows=grid_rows,
        summary=latest_summary,
        task_counts=task_counts_for_js
    )

# --- ★add_or_edit_task ルート (繰り返し対応) ---
@app.route('/add_or_edit_task', methods=['GET', 'POST'])
@app.route('/add_or_edit_task/<int:master_id>', methods=['GET', 'POST'])
@login_required
def add_or_edit_task(master_id=None):
    master_task = db.session.get(MasterTask, master_id) if master_id else None
    if master_task and master_task.user_id != current_user.id:
        flash("権限がありません。")
        return redirect(url_for('todo_list'))

    # Determine the return URL for the 'manage_templates' link
    # date_str をクエリパラメータから取得
    date_str_param = request.args.get('date_str', get_jst_today().strftime('%Y-%m-%d'))
    
    # ★ 編集時/新規作成時で戻り先URLを正しく設定
    if master_id:
        from_url = url_for('add_or_edit_task', master_id=master_id, date_str=date_str_param)
    else:
        from_url = url_for('add_or_edit_task', date_str=date_str_param)


    if request.method == 'POST':
        try:
            # --- Save as Template Logic ---
            # ★ 修正: 'save_as_template' ボタンの value をチェック ( 'true' が送られる)
            if request.form.get('save_as_template') == 'true':
                template_title = request.form.get('master_title', '').strip()
                if not template_title:
                    flash("テンプレートとして保存するには、親タスクのタイトルが必要です。")
                    # ★ 戻り先URL (back_url) を正しく引き継ぐ
                    return redirect(request.args.get('back_url') or from_url)

                # ★★★ 修正点: フォームデータをsessionに保存 ★★★
                session['temp_task_data'] = {
                    'master_title': template_title,
                    'due_date': request.form.get('due_date'),
                    'is_urgent': 'is_urgent' in request.form,
                    'is_habit': 'is_habit' in request.form,
                    'recurrence_type': request.form.get('recurrence_type', 'none'),
                    'recurrence_days': "".join(sorted(request.form.getlist('recurrence_days'))),
                    'subtasks': []
                }
                subtask_count_for_session = 0
                for i in range(1, 21): # Assuming max 20 subtasks from form
                    sub_content = request.form.get(f'sub_content_{i}', '').strip()
                    grid_count_str = request.form.get(f'grid_count_{i}', '0').strip()
                    if sub_content and grid_count_str.isdigit() and int(grid_count_str) > 0:
                        grid_count = int(grid_count_str)
                        session['temp_task_data']['subtasks'].append({'content': sub_content, 'grid_count': grid_count})
                        subtask_count_for_session += 1
                # ★★★ ここまで ★★★

                existing_template = TaskTemplate.query.filter_by(user_id=current_user.id, title=template_title).first()
                if existing_template:
                    template = existing_template
                    SubtaskTemplate.query.filter_by(template_id=template.id).delete()
                    app.logger.info(f"Updating existing template '{template_title}' for user {current_user.id}.")
                else:
                    template = TaskTemplate(title=template_title, user_id=current_user.id)
                    db.session.add(template)
                    db.session.flush() # Need template.id for subtasks
                    app.logger.info(f"Creating new template '{template_title}' for user {current_user.id}.")

                subtask_count = 0
                # ★★★ 修正点: sessionデータからサブタスクを読み込む (すでにあるので流用) ★★★
                for sub in session['temp_task_data']['subtasks']:
                    db.session.add(SubtaskTemplate(template_id=template.id, content=sub['content'], grid_count=sub['grid_count']))
                    subtask_count += 1
                # ★★★ ここまで ★★★

                if subtask_count == 0:
                    flash("サブタスクが1つもないため、テンプレートは保存されませんでした。")
                    db.session.rollback() # Rollback if no subtasks were added
                    # ★★★ 修正点: sessionデータを削除 ★★★
                    session.pop('temp_task_data', None)
                    return redirect(request.args.get('back_url') or from_url)

                db.session.commit()
                flash(f"テンプレート「{template_title}」を保存しました。")
                # manage_templates に戻り先URLを引き継ぐ
                # ★ 修正: manage_templates ではなく、 back_url (元の編集画面) に戻る
                return redirect(request.args.get('back_url') or from_url)

            # --- Save Task Logic ---
            master_title = request.form.get('master_title', '').strip()
            due_date_str = request.form.get('due_date')
            is_urgent = 'is_urgent' in request.form
            is_habit = 'is_habit' in request.form # ★習慣フラグを取得
            recurrence_type = request.form.get('recurrence_type', 'none') # ★繰り返し種別を取得
            recurrence_days = "".join(sorted(request.form.getlist('recurrence_days'))) if recurrence_type == 'weekly' else None # ★繰り返し曜日を取得

            if not master_title:
                flash("親タスクのタイトルを入力してください。")
                return redirect(from_url) # from_url にリダイレクト

            try:
                due_date_obj = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else get_jst_today()
            except ValueError:
                flash("日付の形式が正しくありません。")
                return redirect(from_url)

            if master_task: # Edit existing task
                master_task.title = master_title
                master_task.due_date = due_date_obj # This is start_date for recurring tasks
                master_task.is_urgent = is_urgent
                master_task.is_habit = is_habit
                master_task.recurrence_type = recurrence_type
                master_task.recurrence_days = recurrence_days if recurrence_type == 'weekly' else None
                # Editing a recurring task might need careful handling of existing completions,
                # but for now, just update the properties. We also reset last_reset_date
                # to ensure it appears correctly after edits if needed.
                if recurrence_type != 'none':
                    master_task.last_reset_date = None #編集したらリセット日をクリア

                app.logger.info(f"Updating master task ID {master_task.id} for user {current_user.id}.")
                # Delete existing subtasks before adding new ones
                SubTask.query.filter_by(master_id=master_task.id).delete()
            else: # Create new task
                master_task = MasterTask(
                    title=master_title,
                    due_date=due_date_obj, # This is start_date for recurring tasks
                    user_id=current_user.id,
                    is_urgent=is_urgent,
                    is_habit=is_habit,
                    recurrence_type=recurrence_type,
                    recurrence_days=recurrence_days if recurrence_type == 'weekly' else None,
                    last_reset_date=None #新規作成時はリセット日はnull
                )
                db.session.add(master_task)
                db.session.flush() # Need master_task.id for subtasks
                app.logger.info(f"Creating new master task '{master_title}' for user {current_user.id}.")

            # Add subtasks
            subtask_added = False
            for i in range(1, 21):
                sub_content = request.form.get(f'sub_content_{i}', '').strip()
                grid_count_str = request.form.get(f'grid_count_{i}', '0').strip()
                if sub_content and grid_count_str.isdigit():
                    grid_count = int(grid_count_str)
                    if grid_count > 0:
                        db.session.add(SubTask(master_id=master_task.id, content=sub_content, grid_count=grid_count))
                        subtask_added = True

            if not subtask_added:
                flash("少なくとも1つの有効なサブタスクを入力してください。")
                db.session.rollback() # Rollback if no subtasks added
                return redirect(from_url)

            db.session.commit()
            
            # ★ session データをクリア
            session.pop('temp_task_data', None)
            
            flash("タスクを保存しました。")
            # Redirect intelligently: recurring -> today, normal -> its due date
            redirect_date_str = get_jst_today().strftime('%Y-%m-%d') if recurrence_type != 'none' else master_task.due_date.strftime('%Y-%m-%d')
            return redirect(url_for('todo_list', date_str=redirect_date_str))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error saving task for user {current_user.id}: {e}", exc_info=True)
            flash(f"タスクの保存中にエラーが発生しました: {e}")
            return redirect(from_url)

    # --- GET Request Logic ---
    
    # ★★★ 修正点: sessionから一時データを復元 ★★★
    session_data = session.get('temp_task_data', None) # .pop() から .get() に変更
    # ★★★ ここまで ★★★

    default_date = get_jst_today()
    # date_str_param を GET リクエストでも使う
    if date_str_param:
        try:
            default_date = datetime.strptime(date_str_param, '%Y-%m-%d').date()
        except ValueError:
            pass # Ignore invalid date string
    
    # ★★★ 修正点: sessionデータがある場合、default_dateを上書き ★★★
    if session_data and session_data.get('due_date'):
         try:
            default_date = datetime.strptime(session_data['due_date'], '%Y-%m-%d').date()
         except (ValueError, TypeError):
             pass # パース失敗したら元のdefault_dateを使う
    # ★★★ ここまで ★★★


    templates = TaskTemplate.query.filter_by(user_id=current_user.id).all()
    templates_data = {
        t.id: {
            "title": t.title,
            "subtasks": [{"content": s.content, "grid_count": s.grid_count} for s in t.subtask_templates]
        } for t in templates
    }

    subtasks_for_template = []
    if master_task:
        subtasks_for_template = [{"content": sub.content, "grid_count": sub.grid_count} for sub in master_task.subtasks]
    elif session_data: # ★★★ 修正点: master_taskがなくてもsessionデータからサブタスクを復元 ★★★
        subtasks_for_template = session_data.get('subtasks', [])
    # ★★★ ここまで ★★★

    return render_template(
        'edit_task.html',
        master_task=master_task,
        existing_subtasks=subtasks_for_template,
        default_date=default_date,
        templates=templates,
        templates_data=templates_data,
        session_data=session_data # ★★★ 修正点: sessionデータをテンプレートに渡す ★★★
    )

# --- ★complete_subtask_api (繰り返し対応) ---
@app.route('/api/complete_subtask/<int:subtask_id>', methods=['POST'])
@login_required
def complete_subtask_api(subtask_id):
    subtask = db.session.get(SubTask, subtask_id)
    if not subtask:
        return jsonify({'success': False, 'error': 'Subtask not found'}), 404
    if subtask.master_task.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Permission denied'}), 403

    master_task = subtask.master_task
    today = get_jst_today()
    target_date_str = request.json.get('current_date') if request.is_json else None
    target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date() if target_date_str else today

    # --- Toggle completion status ---
    subtask.is_completed = not subtask.is_completed

    # --- Set completion date (handle recurring tasks) ---
    if subtask.is_completed:
        subtask.completion_date = today # Always set today's date on completion
    else:
        # Only reset completion_date to None if it's NOT a recurring task
        # Recurring task completion dates are handled by the reset logic
        if master_task.recurrence_type == 'none':
             subtask.completion_date = None

    try:
        db.session.commit()
        app.logger.info(f"Subtask {subtask_id} completion toggled to {subtask.is_completed} for user {current_user.id}.")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating subtask {subtask_id} completion: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Database error during commit'}), 500

    # --- Update summary ---
    update_summary(current_user.id) # Recalculate summary after completion change

    # --- Recalculate data needed for UI update ---
    # Refresh master_task and its subtasks from DB
    db.session.refresh(master_task)
    # Eager load subtasks for the specific master task again
    master_task = MasterTask.query.options(selectinload(MasterTask.subtasks)).get(master_task.id)


    # Recalculate visible_subtasks and all_completed_today based on target_date
    today_weekday = str(target_date.weekday())
    is_recurring_today = False
    if master_task.recurrence_type != 'none' and master_task.due_date <= target_date:
        if master_task.recurrence_type == 'daily': is_recurring_today = True
        elif master_task.recurrence_type == 'weekly' and master_task.recurrence_days and today_weekday in master_task.recurrence_days: is_recurring_today = True

    if is_recurring_today:
        visible_subtasks = master_task.subtasks
    else: # Normal task or recurring but not today
        visible_subtasks = [st for st in master_task.subtasks if not st.is_completed or st.completion_date == target_date]

    # visible_subtasks を再度ソート
    visible_subtasks.sort(key=lambda x: x.id)

    subtasks_as_dicts = [{"id": st.id, "content": st.content, "is_completed": st.is_completed, "grid_count": st.grid_count} for st in visible_subtasks]
    master_task.visible_subtasks_json = json.dumps(subtasks_as_dicts) # Add this line back
    master_task.all_completed_today = all(st.is_completed for st in visible_subtasks) if visible_subtasks else False

    # Calculate last completion date among *all* subtasks for header display
    all_completed_ever = all(st.is_completed for st in master_task.subtasks)
    if all_completed_ever:
        completion_dates = [st.completion_date for st in master_task.subtasks if st.completion_date]
        master_task.last_completion_date = max(completion_dates) if completion_dates else None
    else:
        master_task.last_completion_date = None


    # Re-render header HTML
    updated_header_html = render_template('_master_task_header.html', master_task=master_task, current_date=target_date) # Pass current_date

    # --- Recalculate grid and summary data (using same logic as todo_list) ---
    all_master_tasks_query = MasterTask.query.options(
        selectinload(MasterTask.subtasks)
    ).filter(
        MasterTask.user_id == current_user.id,
        MasterTask.subtasks.any()
    )
    all_master_tasks = all_master_tasks_query.all()

    daily_tasks_for_template = []
    recurring_tasks_for_template = []

    for mt in all_master_tasks:
        _is_recurring_today = False
        if mt.recurrence_type != 'none' and mt.due_date <= target_date:
            if mt.recurrence_type == 'daily': _is_recurring_today = True
            elif mt.recurrence_type == 'weekly' and mt.recurrence_days and today_weekday in mt.recurrence_days: _is_recurring_today = True

        if _is_recurring_today:
            mt.visible_subtasks = mt.subtasks # Use all subtasks for recurring today
            recurring_tasks_for_template.append(mt)
        elif mt.recurrence_type == 'none' and mt.due_date == target_date:
             mt.visible_subtasks = [st for st in mt.subtasks if not st.is_completed or st.completion_date == target_date]
             if mt.visible_subtasks: # Only add if there are visible subtasks
                 daily_tasks_for_template.append(mt)

    all_subtasks_for_day_grid = []
    for mt in daily_tasks_for_template: all_subtasks_for_day_grid.extend(mt.visible_subtasks)
    for mt in recurring_tasks_for_template: all_subtasks_for_day_grid.extend(mt.subtasks) # ここは全サブタスクでOK

    total_grid_count = sum(sub.grid_count for sub in all_subtasks_for_day_grid)
    completed_grid_count = sum(sub.grid_count for sub in all_subtasks_for_day_grid if sub.is_completed)

    # Fetch latest summary data (already updated)
    latest_summary = DailySummary.query.filter(DailySummary.user_id == current_user.id).order_by(DailySummary.summary_date.desc()).first()
    summary_data = {
        'streak': latest_summary.streak if latest_summary else 0,
        'average_grids': latest_summary.average_grids if latest_summary else 0.0
    }

    return jsonify({
        'success': True,
        'is_completed': subtask.is_completed,
        'total_grid_count': total_grid_count,
        'completed_grid_count': completed_grid_count,
        'summary': summary_data,
        'updated_header_html': updated_header_html,
        'master_task_id': subtask.master_id
    })


# --- ★習慣カレンダー用ルートとAPI ---
@app.route('/habit_calendar')
@login_required
def habit_calendar():
    # シンプルにテンプレートをレンダリングするだけ
    # 実際のデータはJavaScriptがAPI経由で取得する
    return render_template('habit_calendar.html')

@app.route('/api/habit_calendar/<int:year>/<int:month>')
@login_required
def habit_calendar_data(year, month):
    try:
        start_date = date(year, month, 1)
        # Get the end date of the month correctly
        _, last_day = calendar.monthrange(year, month)
        end_date = date(year, month, last_day)

        app.logger.info(f"Fetching habit data for User {current_user.id} between {start_date} and {end_date}")

        # Fetch completion dates and titles for habit tasks within the month
        # distinct を使って、日付とタイトルが同じ組み合わせは1つだけ取得
        completed_habits = db.session.query(
            SubTask.completion_date,
            MasterTask.title
        ).join(MasterTask).filter(
            MasterTask.user_id == current_user.id,
            MasterTask.is_habit == True, # Only habits
            SubTask.is_completed == True,
            SubTask.completion_date >= start_date,
            SubTask.completion_date <= end_date
        ).distinct(SubTask.completion_date, MasterTask.title).order_by(SubTask.completion_date).all()


        # Group by date and prepare initials/colors
        habits_by_date = {}
        # 色のリストを定義 (少し落ち着いた色合いに変更も検討可)
        colors = ['#EF4444', '#FCD34D', '#10B981', '#3B82F6', '#A855F7', '#EC4899'] # Red, Yellow, Green, Blue, Purple, Pink
        habit_colors = {}
        color_index = 0

        for completion_date, title in completed_habits:
            if completion_date is None: continue

            date_str = completion_date.isoformat()
            initial = title[0].upper() if title else '?'

            # Assign color per habit title
            if title not in habit_colors:
                habit_colors[title] = colors[color_index % len(colors)]
                color_index += 1
            color = habit_colors[title]

            if date_str not in habits_by_date:
                habits_by_date[date_str] = []

            # Add if not already present for that day (日付とタイトルで distinct しているので不要かも)
            # if not any(item['initial'] == initial and item['color'] == color for item in habits_by_date[date_str]):
            habits_by_date[date_str].append({'initial': initial, 'color': color, 'title': title})

        app.logger.info(f"Found {len(habits_by_date)} dates with completed habits for {year}-{month}.")
        return jsonify(habits_by_date)

    except ValueError:
        app.logger.error(f"Invalid year/month requested: {year}-{month}")
        return jsonify({"error": "Invalid date format"}), 400
    except Exception as e:
        app.logger.error(f"Error fetching habit calendar data: {e}", exc_info=True)
        return jsonify({"error": "Failed to fetch habit data"}), 500

@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or not file.filename.endswith('.xlsx'):
            flash('無効なファイル形式です。', "warning") # カテゴリを warning に変更
            return redirect(url_for('import_excel'))
        try:
            # --- ローディング表示のため、すぐにレスポンスを返す代わりに処理を進める ---
            # (ただし、長時間かかる処理はバックグラウンドタスク化が望ましい)
            app.logger.info(f"Starting Excel import for user {current_user.username}...")
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            header = [str(cell.value or '').strip() for cell in sheet[1]] # Noneを空文字に

            # Try to determine column mapping
            col_map = {}
            # Prefer header names if they exist
            expected_headers = {
                'title': ['主タスク', '親タスクのタイトル'],
                'due_date': ['期限日'],
                'sub_content': ['サブタスク内容'],
                'grid_count': ['マス数'],
            }
            found_all = True
            for key, names in expected_headers.items():
                found = False
                for name in names:
                    if name in header:
                        col_map[key] = header.index(name)
                        found = True
                        break
                if not found:
                    found_all = False
                    break

            # Fallback to column indices if headers are missing or different
            if not found_all and len(header) >= 4:
                col_map = {'title': 0, 'due_date': 1, 'sub_content': 2, 'grid_count': 3}
                flash('ヘッダーが見つからないか、形式が異なります。列の順序(A=タイトル, B=期限日, C=サブ内容, D=マス数)でインポートを試みます。', 'info')
            elif not found_all:
                flash('Excelファイルの列数が不足しているか、必要なヘッダーが見つかりません。')
                return redirect(url_for('import_excel'))


            master_tasks_cache = {}
            master_task_count = 0
            sub_task_count = 0
            skipped_rows = 0

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Ensure row has enough columns based on map
                if len(row) <= max(col_map.values()):
                    skipped_rows += 1
                    app.logger.warning(f"Skipping row {row_idx}: Not enough columns.")
                    continue

                master_title = str(row[col_map['title']]).strip() if row[col_map['title']] else None
                due_date_val = row[col_map['due_date']]
                sub_content = str(row[col_map['sub_content']]).strip() if row[col_map['sub_content']] else None
                grid_count_val = row[col_map['grid_count']]

                if not master_title or not sub_content:
                    skipped_rows += 1
                    app.logger.warning(f"Skipping row {row_idx}: Missing master title or subtask content.")
                    continue

                # Process due_date carefully
                due_date = get_jst_today() # Default to today
                if isinstance(due_date_val, datetime):
                    due_date = due_date_val.date()
                elif isinstance(due_date_val, date):
                    due_date = due_date_val
                elif isinstance(due_date_val, (str, int, float)):
                    try:
                        # Attempt to parse common string formats or Excel date numbers
                        if isinstance(due_date_val, (int, float)): # Excel date number
                            # This conversion might need adjustment based on Excel epoch
                            delta = timedelta(days=due_date_val - 25569) # Adjust for Excel epoch difference from Unix epoch
                            due_date = date(1970, 1, 1) + delta
                        else: # String
                            str_date = str(due_date_val).split(" ")[0] # Handle 'YYYY-MM-DD HH:MM:SS'
                            due_date = datetime.strptime(str_date, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        app.logger.warning(f"Could not parse date '{due_date_val}' in row {row_idx}. Using today's date.")
                        # Keep default date (today)

                # Process grid_count
                grid_count = 1 # Default to 1
                if grid_count_val:
                    try:
                        parsed_count = int(float(str(grid_count_val))) # Handle numbers stored as text, floats
                        if parsed_count > 0:
                            grid_count = parsed_count
                    except (ValueError, TypeError):
                        app.logger.warning(f"Could not parse grid count '{grid_count_val}' in row {row_idx}. Using default 1.")


                cache_key = (master_title, due_date)
                if cache_key not in master_tasks_cache:
                    # Import as non-recurring task by default
                    master_task = MasterTask(title=master_title, due_date=due_date, user_id=current_user.id, recurrence_type='none')
                    db.session.add(master_task)
                    db.session.flush() # Get the ID for the cache
                    master_tasks_cache[cache_key] = master_task
                    master_task_count += 1
                else:
                    master_task = master_tasks_cache[cache_key]

                db.session.add(SubTask(master_id=master_task.id, content=sub_content, grid_count=grid_count))
                sub_task_count += 1

            db.session.commit()
            app.logger.info(f"Excel import complete for user {current_user.username}. Imported {master_task_count} master tasks, {sub_task_count} subtasks. Skipped {skipped_rows} rows.")
            flash(f'{master_task_count}件の親タスク ({sub_task_count}件のサブタスク) をインポートしました。{skipped_rows}行はスキップされました。')
            return redirect(url_for('todo_list'))

        except Exception as e:
            db.session.rollback() # Rollback in case of any error
            app.logger.error(f'Excel import failed for user {current_user.username}: {e}', exc_info=True)
            flash(f'インポート処理中にエラーが発生しました: {e}')
            return redirect(url_for('import_excel'))

    return render_template('import.html')

# --- 他のルート (manage_templates, delete_template, scratchpad, export_scratchpad, export_to_sheet, admin routes...) ---
@app.route('/templates', methods=['GET', 'POST'])
@login_required
def manage_templates():
    back_url = request.args.get('back_url') # Get potential back URL

    if request.method == 'POST':
        # ★ テンプレート保存ロジックは add_or_edit_task に移動
        flash("テンプレートの保存方法が変更されました。")
        return redirect(url_for('manage_templates', back_url=back_url))

    # GET request
    templates = TaskTemplate.query.filter_by(user_id=current_user.id).order_by(TaskTemplate.title).all()
    return render_template('manage_templates.html', templates=templates, back_url=back_url)


@app.route('/delete_template/<int:template_id>', methods=['POST'])
@login_required
def delete_template(template_id):
    template = db.session.get(TaskTemplate, template_id)
    if not template:
        flash("テンプレートが見つかりません。")
        return redirect(url_for('manage_templates'))
    if template.user_id != current_user.id:
        flash("権限がありません。")
        return redirect(url_for('manage_templates'))

    try:
        title = template.title
        # Subtasks are deleted automatically due to cascade rule
        db.session.delete(template)
        db.session.commit()
        flash(f"テンプレート「{title}」を削除しました。")
        app.logger.info(f"Deleted template '{title}' (ID: {template_id}) for user {current_user.id}.")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting template {template_id} for user {current_user.id}: {e}", exc_info=True)
        flash(f"テンプレートの削除中にエラーが発生しました: {e}")

    # Preserve back_url if present in referrer or args
    back_url = request.args.get('back_url') or request.referrer or url_for('manage_templates')
    return redirect(back_url)


# --- 8. スクラッチパッド関連ルート ---
@app.route('/scratchpad')
@login_required
def scratchpad():
    # Render the simple scratchpad page
    return render_template('scratchpad.html')


@app.route('/export_scratchpad', methods=['POST'])
@login_required
def export_scratchpad():
    # Check if request is JSON
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Invalid request format.'}), 400

    tasks_to_add = request.json.get('tasks')
    if not tasks_to_add or not isinstance(tasks_to_add, list):
        return jsonify({'success': False, 'message': 'No valid tasks provided.'}), 400

    today = get_jst_today()
    master_title = f"{today.strftime('%Y-%m-%d')}のクイックタスク"

    try:
        # Find or create the master task for today's quick tasks
        master_task = MasterTask.query.filter_by(user_id=current_user.id, title=master_title, due_date=today, recurrence_type='none').first()
        if not master_task:
            master_task = MasterTask(title=master_title, due_date=today, user_id=current_user.id, recurrence_type='none')
            db.session.add(master_task)
            db.session.flush() # Need the ID
            app.logger.info(f"Created quick task master '{master_title}' for user {current_user.id}.")

        # Add each scratchpad item as a subtask
        added_count = 0
        for task_content in tasks_to_add:
            if isinstance(task_content, str) and task_content.strip():
                db.session.add(SubTask(master_id=master_task.id, content=task_content.strip(), grid_count=1))
                added_count += 1

        if added_count > 0:
            db.session.commit()
            app.logger.info(f"Exported {added_count} scratchpad tasks to quick tasks for user {current_user.id}.")
            return jsonify({'success': True, 'message': f'{added_count}件のタスクを追加しました。'})
        else:
            app.logger.info(f"No valid tasks found in scratchpad export request for user {current_user.id}.")
            return jsonify({'success': False, 'message': '追加する有効なタスクがありませんでした。'})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error exporting scratchpad tasks for user {current_user.id}: {e}", exc_info=True)
        return jsonify({'success': False, 'message': 'タスクの追加中にエラーが発生しました。'}), 500


# --- スプレッドシート関連 ---
def get_gspread_client():
    sa_info = os.environ.get('GSPREAD_SERVICE_ACCOUNT')
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

    try:
        if sa_info:
            sa_creds = json.loads(sa_info)
            creds = ServiceAccountCredentials.from_json_keyfile_dict(sa_creds, scope)
            app.logger.info("Authenticating GSpread using environment variable.")
        else:
            # Fallback to local file if env var is not set
            creds = ServiceAccountCredentials.from_json_keyfile_name('service_account.json', scope)
            app.logger.info("Authenticating GSpread using service_account.json file.")
        return gspread.authorize(creds)
    except FileNotFoundError:
        app.logger.error("GSpread authentication failed: service_account.json not found and GSPREAD_SERVICE_ACCOUNT env var not set.")
        return None
    except Exception as e:
        app.logger.error(f"GSpread authentication failed: {e}", exc_info=True)
        return None

@app.route('/export_to_sheet', methods=['POST'])
@login_required
def export_to_sheet():
    if not current_user.spreadsheet_url:
        flash("スプレッドシートURLが設定されていません。設定ページでURLを登録してください。")
        return redirect(url_for('settings')) # Redirect to settings if URL is missing

    # Find completed, non-recurring tasks that have a completion date
    completed_tasks = SubTask.query.join(MasterTask).filter(
        MasterTask.user_id == current_user.id,
        MasterTask.recurrence_type == 'none', # Export only non-recurring tasks
        SubTask.is_completed == True,
        SubTask.completion_date != None
    ).order_by(SubTask.completion_date).all()

    if not completed_tasks:
        flash("スプレッドシートに書き出す完了済みタスクがありません。")
        return redirect(url_for('todo_list'))

    gc = get_gspread_client()
    if not gc:
        flash("スプレッドシート認証に失敗しました。管理者設定またはサービスアカウントファイルを確認してください。")
        return redirect(url_for('settings')) # Redirect to settings on auth failure

    try:
        app.logger.info(f"Attempting to open spreadsheet: {current_user.spreadsheet_url}")
        sh = gc.open_by_url(current_user.spreadsheet_url)
        worksheet = sh.sheet1 # Assume first sheet

        header = ['主タスクID', '主タスク', 'サブタスク内容', 'マス数', '期限日', '完了日', '遅れた日数']
        existing_header = worksheet.row_values(1)

        # Ensure header exists, create if not
        if not existing_header or existing_header != header:
             # Check if sheet is empty or has different header
            if not existing_header and worksheet.row_count == 0:
                worksheet.append_row(header)
                app.logger.info("Appended header row to empty spreadsheet.")
            elif existing_header != header :
                # Decide how to handle existing different header? Overwrite? Append anyway? Error?
                # For now, let's append anyway but log a warning
                app.logger.warning("Spreadsheet header does not match expected format. Appending data anyway.")

        # Fetch existing records efficiently only if needed to prevent duplicates
        app.logger.info("Fetching existing records from spreadsheet...")
        try:
            records = worksheet.get_all_values() # Can be slow for large sheets
            existing_records = records[1:] if len(records) > 1 else []
            # Create a set of unique keys (e.g., master_title, sub_content, completion_date)
            # Adjust indices based on actual sheet columns
            existing_keys = set( (rec[1], rec[2], rec[5]) for rec in existing_records if len(rec) >= 6)
            app.logger.info(f"Found {len(existing_keys)} existing unique keys.")
        except gspread.exceptions.APIError as api_err:
            # Handle potential API errors like rate limits
            app.logger.error(f"GSpread API Error fetching records: {api_err}")
            flash(f"スプレッドシートからのデータ取得中にエラーが発生しました: {api_err}")
            return redirect(url_for('todo_list'))


        data_to_append = []
        app.logger.info(f"Processing {len(completed_tasks)} completed tasks for export...")
        for subtask in completed_tasks:
            # completion_date should not be None due to query filter, but check anyway
            if not subtask.completion_date: continue

            completion_date_str = subtask.completion_date.strftime('%Y-%m-%d')
            due_date_str = subtask.master_task.due_date.strftime('%Y-%m-%d')
            key = (subtask.master_task.title, subtask.content, completion_date_str)

            if key not in existing_keys:
                day_diff = (subtask.completion_date - subtask.master_task.due_date).days
                data_to_append.append([
                    subtask.master_task.id, subtask.master_task.title, subtask.content,
                    subtask.grid_count, due_date_str,
                    completion_date_str, day_diff
                ])
                existing_keys.add(key) # Add new key to prevent duplicates within the same batch

        if data_to_append:
            app.logger.info(f"Appending {len(data_to_append)} new rows to spreadsheet...")
            worksheet.append_rows(data_to_append, value_input_option='USER_ENTERED')
            flash(f"{len(data_to_append)}件の新しい完了タスクをスプレッドシートに書き出しました。")
        else:
            flash("スプレッドシートに書き出す新しい完了タスクはありませんでした。")

    except gspread.exceptions.SpreadsheetNotFound:
        app.logger.error(f"Spreadsheet not found for URL: {current_user.spreadsheet_url}")
        flash("指定されたURLのスプレッドシートが見つかりません。URLを確認し、共有設定を確認してください。")
        return redirect(url_for('settings'))
    except gspread.exceptions.APIError as api_err:
        app.logger.error(f"GSpread API Error during export: {api_err}")
        flash(f"スプレッドシートへの書き込み中にAPIエラーが発生しました: {api_err}")
        return redirect(url_for('settings')) # Redirect to settings might be helpful
    except Exception as e:
        app.logger.error(f"Unexpected error during spreadsheet export: {e}", exc_info=True)
        flash(f"スプレッドシートへの書き込み中に予期せぬエラーが発生しました: {e}")
        return redirect(url_for('todo_list'))

    return redirect(url_for('todo_list'))

# --- 10. 管理者用ルート ---
@app.route('/admin')
@login_required
def admin_panel():
    if not current_user.is_admin:
        flash("管理者権限がありません。")
        return redirect(url_for('todo_list'))
    try:
        users = User.query.order_by(User.id).all()
        return render_template('admin.html', users=users)
    except Exception as e:
        app.logger.error(f"Error loading admin panel: {e}", exc_info=True)
        flash("ユーザーリストの読み込み中にエラーが発生しました。")
        return redirect(url_for('todo_list'))


@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        flash("管理者権限がありません。")
        return redirect(url_for('todo_list'))
    if user_id == current_user.id:
        flash("自分自身のアカウントは削除できません。")
        return redirect(url_for('admin_panel'))

    user_to_delete = db.session.get(User, user_id)
    if not user_to_delete:
        flash("指定されたユーザーが見つかりません。")
        return redirect(url_for('admin_panel'))

    try:
        username = user_to_delete.username
        # Cascade should handle related data deletion
        db.session.delete(user_to_delete)
        db.session.commit()
        flash(f"ユーザー「{username}」とその関連データを削除しました。")
        app.logger.info(f"Admin {current_user.username} deleted user {username} (ID: {user_id}).")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting user {user_id} by admin {current_user.username}: {e}", exc_info=True)
        flash(f"ユーザーの削除中にエラーが発生しました: {e}")

    return redirect(url_for('admin_panel'))


@app.route('/admin/reset_password/<int:user_id>', methods=['POST'])
@login_required
def reset_password(user_id):
    if not current_user.is_admin:
        flash("管理者権限がありません。")
        return redirect(url_for('admin_panel'))

    user_to_reset = db.session.get(User, user_id)
    if not user_to_reset:
        flash("指定されたユーザーが見つかりません。")
        return redirect(url_for('admin_panel'))
    if user_to_reset.is_admin:
        flash("管理者ユーザーのパスワードはこの方法ではリセットできません。")
        return redirect(url_for('admin_panel'))

    try:
        new_password = secrets.token_hex(8) # Generate an 8-byte (16 hex chars) random password
        user_to_reset.password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        user_to_reset.password_reset_required = True # Force user to change password on next login
        db.session.commit()

        # Flash the new password for the admin to copy
        flash(f"ユーザー「{user_to_reset.username}」の新しい一時パスワードは「{new_password}」です。コピーしてユーザーに伝えてください。次回のログイン時にパスワード変更が要求されます。", 'success')
        app.logger.info(f"Admin {current_user.username} reset password for user {user_to_reset.username} (ID: {user_id}).")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error resetting password for user {user_id} by admin {current_user.username}: {e}", exc_info=True)
        flash(f"パスワードのリセット中にエラーが発生しました: {e}")

    return redirect(url_for('admin_panel'))


@app.route('/admin/export_user_data/<int:user_id>', methods=['POST'])
@login_required
def export_user_data(user_id):
    if not current_user.is_admin:
        flash("管理者権限がありません。")
        return redirect(url_for('admin_panel'))

    user = db.session.get(User, user_id)
    if not user:
        flash("指定されたユーザーが見つかりません。")
        return redirect(url_for('admin_panel'))

    try:
        # Fetch all tasks for the user, ordered logically
        all_tasks = SubTask.query.join(MasterTask).filter(
            MasterTask.user_id == user.id
        ).options(
            selectinload(SubTask.master_task) # Eager load master task details
        ).order_by(
            MasterTask.due_date, MasterTask.id, SubTask.id
        ).all()


        if not all_tasks:
            flash(f"ユーザー「{user.username}」には書き出すタスクがありません。")
            return redirect(url_for('admin_panel'))

        app.logger.info(f"Starting data export for user {user.username} (ID: {user_id}) by admin {current_user.username}. Found {len(all_tasks)} subtasks.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{user.username}_tasks"

        # Define header
        header = [
            '親タスクID', '親タスクタイトル', '期限日/開始日', '緊急', '習慣',
            '繰り返し種別', '繰り返し曜日',
            'サブタスクID', 'サブタスク内容', 'マス数', '完了状態', '完了日', '遅れた日数'
        ]
        ws.append(header)

        # Append data rows
        for subtask in all_tasks:
            master = subtask.master_task
            completion_date_str = subtask.completion_date.strftime('%Y-%m-%d') if subtask.completion_date else ''
            day_diff = (subtask.completion_date - master.due_date).days if subtask.completion_date and master.due_date and master.recurrence_type == 'none' else '' # 通常タスクのみ遅延計算

            ws.append([
                master.id, master.title, master.due_date.strftime('%Y-%m-%d'),
                'Yes' if master.is_urgent else 'No',
                'Yes' if master.is_habit else 'No',
                master.recurrence_type,
                master.recurrence_days or '',
                subtask.id, subtask.content, subtask.grid_count,
                '完了' if subtask.is_completed else '未完了',
                completion_date_str,
                day_diff
            ])

        # Save workbook to a BytesIO buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'{user.username}_all_tasks_{get_jst_today().strftime("%Y%m%d")}.xlsx'
        app.logger.info(f"Successfully generated export file: {filename}")

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        app.logger.error(f"Error exporting data for user {user_id} by admin {current_user.username}: {e}", exc_info=True)
        flash(f"ユーザーデータの書き出し中にエラーが発生しました: {e}")
        return redirect(url_for('admin_panel'))


# --- PWA対応: オフライン同期API ---
@app.route('/api/sync', methods=['POST'])
@login_required
def sync_api():
    if not request.is_json:
        return jsonify({"success": False, "error": "Invalid JSON"}), 400

    data = request.json
    user_id = current_user.id
    today = get_jst_today()

    try:
        # 1. Process new tasks
        new_tasks = data.get('new_tasks', [])
        app.logger.info(f"Sync: Processing {len(new_tasks)} new tasks for user {user_id}")
        for task_data in new_tasks:
            title = task_data.get('title')
            due_date_str = task_data.get('due_date')
            subtasks = task_data.get('subtasks')
            if not title or not due_date_str or not subtasks:
                app.logger.warning(f"Skipping incomplete new task data: {task_data}")
                continue

            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            master_task = MasterTask(
                user_id=user_id, title=title, due_date=due_date,
                is_urgent=task_data.get('is_urgent', False),
                is_habit=task_data.get('is_habit', False),
                recurrence_type=task_data.get('recurrence_type', 'none'),
                recurrence_days=task_data.get('recurrence_days')
            )
            db.session.add(master_task)
            db.session.flush()

            for sub_data in subtasks:
                if sub_data.get('content') and sub_data.get('grid_count', 0) > 0:
                    db.session.add(SubTask(master_id=master_task.id, content=sub_data['content'], grid_count=sub_data['grid_count']))

        # 2. Process scratchpad tasks
        scratchpad_tasks = data.get('scratchpad_tasks', [])
        app.logger.info(f"Sync: Processing {len(scratchpad_tasks)} scratchpad tasks for user {user_id}")
        if scratchpad_tasks:
            master_title = f"{today.strftime('%Y-%m-%d')}のクイックタスク"
            master_task = MasterTask.query.filter_by(user_id=user_id, title=master_title, due_date=today, recurrence_type='none').first()
            if not master_task:
                master_task = MasterTask(title=master_title, due_date=today, user_id=user_id, recurrence_type='none')
                db.session.add(master_task)
                db.session.flush()
            for task_content in scratchpad_tasks:
                if isinstance(task_content, str) and task_content.strip():
                    db.session.add(SubTask(master_id=master_task.id, content=task_content.strip(), grid_count=1))

        # 3. Process new templates
        new_templates = data.get('new_templates', [])
        app.logger.info(f"Sync: Processing {len(new_templates)} new templates for user {user_id}")
        for template_data in new_templates:
            title = template_data.get('title')
            subtasks = template_data.get('subtasks')
            if not title or not subtasks:
                app.logger.warning(f"Skipping incomplete template data: {template_data}")
                continue

            existing_template = TaskTemplate.query.filter_by(user_id=user_id, title=title).first()
            if existing_template:
                template = existing_template
                SubtaskTemplate.query.filter_by(template_id=template.id).delete()
            else:
                template = TaskTemplate(title=title, user_id=user_id)
                db.session.add(template)
                db.session.flush()

            for sub_data in subtasks:
                if sub_data.get('content') and sub_data.get('grid_count', 0) > 0:
                    db.session.add(SubtaskTemplate(template_id=template.id, content=sub_data['content'], grid_count=sub_data['grid_count']))

        # 4. Process completed tasks
        completed_tasks = data.get('completed_tasks', [])
        app.logger.info(f"Sync: Processing {len(completed_tasks)} completed task updates for user {user_id}")
        for comp_data in completed_tasks:
            subtask_id = comp_data.get('subtaskId')
            is_completed = comp_data.get('isCompleted')
            if subtask_id is None: continue

            subtask = db.session.get(SubTask, subtask_id)
            if subtask and subtask.master_task.user_id == user_id:
                subtask.is_completed = is_completed
                # 完了日は今日の日付を設定 (オフライン時の日付ではなく同期時の日付)
                # 繰り返しでないタスクの未完了時はNoneに
                if is_completed:
                    subtask.completion_date = today
                elif subtask.master_task.recurrence_type == 'none':
                     subtask.completion_date = None
                # (繰り返しタスクはリセット時にNoneになる)

        db.session.commit()
        app.logger.info(f"Successfully synced offline data for user {user_id}")
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error during offline sync for user {user_id}: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Sync failed on server"}), 500

# --- 11. アプリの実行 ---
if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    # Automatically create tables if run directly (useful for local dev)
    with app.app_context():
        try:
            db.create_all() # This will now add the new columns
            app.logger.info("Database tables checked/created/updated.")
            # Set admin flag if specified in env vars
            admin_username = os.environ.get('ADMIN_USERNAME')
            if admin_username:
                admin_user = User.query.filter_by(username=admin_username).first()
                if admin_user and not admin_user.is_admin:
                    admin_user.is_admin = True
                    db.session.commit()
                    app.logger.info(f"User '{admin_username}' set as admin.")
        except Exception as e:
            app.logger.error(f"Error during initial DB setup/admin check: {e}", exc_info=True)

    # Use Gunicorn for production, Flask dev server for debug
    if os.environ.get("FLASK_ENV") == "production":
        app.logger.info(f"Starting Gunicorn on port {port}.")
        # Gunicorn is typically started via Procfile, not here directly
        pass # Gunicorn runs externally
    else:
        app.logger.info(f"Starting Flask development server on port {port}.")
        # Enable reloader and debugger for development
        app.run(debug=True, host='0.0.0.0', port=port)
